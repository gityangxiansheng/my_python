import os  
import openpyxl  
import pandas as pd  

# 创建一个空字典来存储数据  
data_dict = {}  

variable4 = ""  # 用于获取字典的键  
var3 = []  # 用于储存位置信息  
  
# 创建变量1，用于储存未匹配到项的文件名  
no_match_files = []  
  
# 创建变量2，用于储存位置信息  
location_info = []  
  
# 指定文件夹路径（将“your_folder_path”替换为实际的文件夹路径）  
folder_path = r"C:\Users\Administrator\Desktop\新建文件夹 (5)"  
  
# 如果文件夹路径存在，则遍历文件夹内的所有文件  
if os.path.exists(folder_path):  
    for filename in os.listdir(folder_path):  
        # 检查文件是否为Excel文件  
        if filename.endswith(".xlsx"):  
            # 打开Excel表格，并提取当前Excel表格文件名称  
            workbook = openpyxl.load_workbook(os.path.join(folder_path, filename))  
            sheet = workbook.active  
            table_name = filename[:-5]  # 提取表格名称，假设文件名格式为"table_name.xlsx"  
            key = table_name[:3].lower()  # 使用lower()确保不区分大小写  
              
            # 搜索Excel表格内与Excel表格文件名前三个字符相同的单元格  
            matching_cell = None  
            for row in sheet.iter_rows():  
                for cell in row:  
                    if cell.value and str(cell.value)[:3].lower() == key:  
                        matching_cell = cell  
                        break  
              
            # 如果找到与文件名前三个字符相同的单元格，则执行以下操作：  
            if matching_cell:  
                data_dict[key] = {}  # 确保字典中存在该键  
                data_dict[key][matching_cell.value] = sheet.cell(row=matching_cell.row + 1, column=matching_cell.column).value  # 更新或添加新的值  
                location_info.append((matching_cell.row, matching_cell.column))  # 将位置信息添加到变量2中  
                # 通过坐标信息定位到符合条件的单元格右侧的第一个单元格的位置信息并更新变量2的值  
                location_info.append((matching_cell.row, matching_cell.column + 1))  # 将位置信息添加到变量2中  
                data_dict[key][matching_cell.value] = sheet.cell(row=location_info[-1][0], column=location_info[-1][1]).value  # 更新字典的值 # 更新字典的值  
            else:  # 如果未找到与文件名前三个字符相同的单元格，则将当前表格文件名称添加到变量1中  
                no_match_files.append(filename)  
else:  # 如果文件夹路径不存在，则打印错误信息  
    print("指定的文件夹路径不存在！")  

# 读取Excel表格内容到DataFrame  
file_path = "C:\\Users\\Administrator\\Desktop\\新建文件夹 (7)\\开展情况日报表2024.1.18.xlsx"  
df = pd.read_excel(file_path, engine='openpyxl')  # 使用openpyxl引擎读取Excel文件  

# variable4 = [key for key in data_dict]
# print("表格内容："+str(variable4))

# 遍历DataFrame的每一行来搜索匹配项  
for index, row in df.iterrows():  
    for column in df.columns:  
        # 检查单元格的值是否与变量4匹配
        
        if row[column] == str(variable4) [:3]:  # 这里使用in运算符检查值是否在字典中  
            # 将符合条件的单元格的坐标储存至变量3  
            var3 = (index, column)  # 获取行和列的位置索引  
            break  # 找到后退出循环  
    if var3:  # 如果变量3已经有值，说明找到了匹配的单元格  
        # 更新变量3的坐标信息定位到匹配单元格的下一列同一行的单元格的位置信息  
        next_column_index = (var3[1] + 1) % len(df.columns)  # 计算下一列的位置索引，使用模运算确保索引不会越界  
        var3 = (var3[0], next_column_index)  # 更新坐标为下一列的位置索引  
        # 在变量3坐标信息对应的Excel表格的单元格内填充将变量4作为字典的键对应的值（这里需要根据实际情况填充）  
        df.at[var3[0], df.columns[var3[1]]] = data_dict[variable4]  # 这里需要您提供实际的填充值，假设它来自data_dict字典的值  
        break  # 填充后退出循环  
  
# 将DataFrame的内容写回到Excel表格中（覆盖原有内容）  
df.to_excel(file_path, index=False, engine='openpyxl')



  
# # 打印字典、变量1和变量2的值（注意：变量2是一个位置信息的列表）  
# print("字典内容：", data_dict[key])  
# print("没有匹配项的Excel文件：", no_match_files)  
# print("位置信息：", location_info)